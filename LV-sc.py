from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import pandas as pd
from openpyxl import load_workbook, Workbook
import time
from os import path

# Logo User Scraping #https://de.lovoo.com/
# 1- Loginbutton.click()
# 2- send keys: Email-Adresse
# 3- send keys: PW
# 4- Loginbutton.click()
# 5- Matchbutton.click()   # Ab hier kommt ein loop
# 6- Profil.click()
# 7- Steckbrief.click() # BIS HIERHIN BIN ICH GEKOMMEN. (NO SUCH ELEMENT ERROR)
# 8- df.to_excel Data
# 9- driver.back
# 10- Match.button.click()
# 6- Profil.click()
# 7- Steckbrief.click()
# 8- df.to_excel Data
# 9- driver.back

class LovooScraping():

    def __init__(self):
        self.driver= webdriver.Chrome(executable_path='/Users/Fabi/Downloads/chromedriver')
        self.df = pd.DataFrame()
        self.name_list = []
        self.wohnort_list = []
        self.interessiert_list = []
        self.aussehen_list = []
        self.raucher_list = []
        self.wohnsituation_list = []


    def compile_data(self):

        # Name
        name = self.driver.find_elements_by_css_selector('#profile-details > div > div.h5.relative.bg-gender.text-white.padding-lg.text-bold.no-space-after.no-space-before > span')
        self.name_list = [value.text for value in name]
        print(self.name_list)

        # Wohnort
        wohnort = self.driver.find_elements_by_css_selector('#profile-details > div > div.space-before-sm > div > div.tab-pane.active > div > div:nth-child(1) > div > dl > dd')
        self.wohnort_list = [value.text for value in wohnort]
        print(self.wohnort_list)

        # Interessiert an
        interessiert = self.driver.find_elements_by_css_selector('#profile-details > div > div.space-before-sm > div > div.tab-pane.active > div > div:nth-child(3) > div > dl > dd')
        self.interessiert_list = [value.text for value in interessiert]
        print(self.interessiert_list)

        # Aussehen
        aussehen = self.driver.find_elements_by_css_selector('#profile-details > div > div.space-before-sm > div > div.tab-pane.active > div > div:nth-child(5) > div > dl > dd')
        self.aussehen_list = [value.text for value in aussehen]
        print(self.aussehen_list)

        # Raucher
        raucher = self.driver.find_elements_by_css_selector('#profile-details > div > div.space-before-sm > div > div.tab-pane.active > div > div:nth-child(7) > div > dl > dd')
        self.raucher_list = [value.text for value in raucher]
        print(self.raucher_list)

        # Wohnsituation
        wohnsituation = self.driver.find_elements_by_css_selector('#profile-details > div > div.space-before-sm > div > div.tab-pane.active > div > div:nth-child(9) > div > dl > dd')
        self.wohnsituation_list = [value.text for value in wohnsituation]


        for i, name in enumerate(self.name_list):
            try:
                self.df.loc[i, "name"] = name

            except Exception as e:

                print('Name konnte nicht gefunden werden', e.message)

            try:
                self.df.loc[i, "Wohnort"] = self.wohnort_list[i]

            except Exception as e:

                print('Wohnort konnte nicht gefunden werden', e.message)
            try:
                self.df.loc[i, "Aussehen"] = self.aussehen_list[i]

            except Exception as e:

                print('Aussehen konnte nicht gefunden werden', e.message)

            try:
                self.df.loc[i, "Raucher"] = self.raucher_list[i]

            except Exception as e:

                print('Raucher konnte nicht gefunden werden', e.message)

            try:
                self.df.loc[i, "Wohnsituation"] = self.wohnsituation_list[i]

            except Exception as e:

                print('Wohnsituation konnte nicht gefunden werden', e.message)

        print("Excel sheet created!")

# if path.isfile(wb_name):
#     wb = load_workbook(wb_name)
# else:
#     wb = Workbook()
#
# try:
#     sh = wb[sh_name]
#
# except KeyError:
#     sh = wb.create_sheet(sh_name)

# header = ['Name', 'Wohnort', 'Aussehen', 'Raucher', 'Wohnsituation']

# for j, elem in enumerate(header):
#     sh.cell(row=1, column=j+1, value=elem)

driver = LovooScraping()  # // steht jetzt im for-loop

link = 'https://de.lovoo.com/'

driver.driver.get(link)

einloggen_button = driver.driver.find_element_by_xpath("/html/body/div[1]/div/div[3]/button[2]")
einloggen_button.click()

username = driver.driver.find_element_by_xpath("//*[@id='form']/div[1]/input")
password = driver.driver.find_element_by_xpath("//*[@id='form']/div[2]/div[1]/input")

username.send_keys("damien9445@gmail.com")
time.sleep(5)
password.send_keys("xae1a5bb")
time.sleep(5)
driver.driver.find_element_by_xpath("//*[@id='form']/div[2]/div[2]/button").click()
# if einloggen_button.click()
time.sleep(5)

wb_name = "Lovoo.xlsx"
sh_name = 'Lovoo'
wb = load_workbook(wb_name)
sh = wb[sh_name]
last_line = sh.max_row+1

for i in range(5):
    driver.driver.find_element_by_xpath("//*[@id='topmenu']/div/nav/div/div[2]/div/ul[1]/li[2]/a").click()
    time.sleep(2)
    driver.driver.find_element_by_xpath("//*[@id='page-content']/match/div/div/div[1]/a/div").click()
    time.sleep(2)
    driver.driver.find_element_by_css_selector('#profile-details > div > div.space-before-sm > ul > li:nth-child(2) > a').click()
    time.sleep(1)
    driver.compile_data()

    print(driver.df.to_dict())
    row = list(driver.df.to_dict().values())

    for j, elem in enumerate(row):
        sh.cell(row=last_line+i, column=j+1, value=elem[0])

    driver.driver.back()
    time.sleep(5)

wb.save(wb_name)

driver.driver.close()

# for i in range(10):
#     driver = LovooScraping()
#     driver.driver.find_element_by_xpath("//*[@id='page-content']/match/div/div/div[1]/a/div").click()
#     time.sleep(5)
#     driver.driver.find_element_by_css_selector('#profile-details > div > div.space-before-sm > ul > li:nth-child(2) > a').click()
#     time.sleep(5)
#     driver.compile_data()
#     driver.back()
#     driver.forward()
#     driver.compile_data()

# driver.df.to_excel(writer, sheet_name='Lovoo')
#
# writer.save()
